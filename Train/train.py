import openpyxl
import config
import math
from sklearn import svm
from sklearn.ensemble import RandomForestClassifier
from sklearn.ensemble import ExtraTreesClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.metrics import accuracy_score
from sklearn.model_selection import train_test_split
from sklearn import cross_validation
from sklearn import metrics
from sklearn import preprocessing
from sklearn.grid_search import GridSearchCV
# from sklearn.cross_validation import cross_val_score
from sklearn.metrics import hamming_loss
from datetime import datetime

configs = config.configs

# 记录程序开始执行时间
start = datetime.now()

# readBook = openpyxl.load_workbook(configs['condensedFeaturesPath'])
# readSheet = readBook.active

readBook = openpyxl.load_workbook('/Users/ming.zhou/NLP/datasets/eliminateParaTag.xlsx')
readSheet = readBook.active

# readTestBook = openpyxl.load_workbook(configs['condensedTestFeaturesPath'])
# readTestSheet = readTestBook.active

readTestBook = openpyxl.load_workbook('/Users/ming.zhou/NLP/datasets/eliminateParaTagTest.xlsx')
readTestSheet = readTestBook.active

readAllBook = openpyxl.load_workbook(configs['allFeaturesPath'])
readAllSheet = readAllBook.active

# 获取训练集的句子特征向量以及句子标签
def getFeatureVectorAndTag():
    featureVectorList = []
    sentenceTag = []
    for i in range(readSheet.max_row - 1):

        featureVector = []
        sentenceTag.append(readSheet['E' + str(i + 2)].value)
        # 先添加段落特征
        featureVector.append(float(readSheet['A' + str(i + 2)].value.strip().split('-')[1]))
        for j in range(readSheet.max_column - 5):
            if isinstance(readSheet.cell(row = i + 2, column = j + 6).value, str):
                featureVector.append(float(readSheet.cell(row=i + 2, column=j + 6).value))
            else:
                featureVector.append(readSheet.cell(row = i + 2, column = j + 6).value)
        featureVectorList.append(featureVector)
    return featureVectorList, sentenceTag

# 获取测试集的句子特征向量以及句子标签
def getTestFeatureVectorAndTag():
    featureVectorList = []
    sentenceTag = []
    for i in range(readTestSheet.max_row - 1):

        featureVector = []
        sentenceTag.append(readTestSheet['E' + str(i + 2)].value)
        # 先添加段落特征
        featureVector.append(float(readTestSheet['A' + str(i + 2)].value.strip().split('-')[1]))
        for j in range(readTestSheet.max_column - 5):
            if isinstance(readTestSheet.cell(row = i + 2, column = j + 6).value, str):
                featureVector.append(float(readTestSheet.cell(row=i + 2, column=j + 6).value))
            else:
                featureVector.append(readTestSheet.cell(row = i + 2, column = j + 6).value)
        featureVectorList.append(featureVector)
    return featureVectorList, sentenceTag

def adjustParameter(X, y, flag='ExtraTrees'):
    clf = ExtraTreesClassifier(min_samples_split=100,
                                 min_samples_leaf=20,max_depth=8,max_features='sqrt' ,random_state=10)
    if flag == 'RandomForest':
        clf = RandomForestClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()
    param_grid = {
        'n_estimators':range(5, 20, 2)
    }
    gridSearch = GridSearchCV(estimator=clf,param_grid=param_grid, scoring='roc_auc', cv = 5)
    gridSearch.fit(X, y)
    print(gridSearch.grid_scores_)
    print(gridSearch.best_params_)
    print(gridSearch.best_score_)

# 训练并测试训练集的准确度
def doTrain(X, Y, flag='RandomForest'):

    clf = RandomForestClassifier(n_estimators=10)
    if flag == 'ExtraTrees':
        clf = ExtraTreesClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    clf.fit(X, Y)
    # clf = clf.fit(X, Y)

    Y_pred = clf.predict(X)
    print(accuracy_score(Y, Y_pred))
    print(Y)
    print(Y_pred)
    print('Y len is ', len(Y),'Y_pred len is ', len(Y_pred))
    # scores = cross_val_score(clf, X, Y)
    # print(scores.mean())

    compare = {}
    j = 0
    for i in range(len(Y)):
        if Y[i] == 5:
            j += 1
        YAndYpred = str(Y[i]) + '-' + str(Y_pred[i])
        # print(YAndYpred)
        if(Y[i] != Y_pred[i]):
            if YAndYpred not in compare:
                compare[YAndYpred] = 1
            else:
                compare[YAndYpred] = compare[YAndYpred] + 1

    print(sorted(compare.items(), key=lambda d: -d[1]))

# 测试测试集的准确度
def doTrainByTestSet(train_X, train_Y, test_X, test_Y, flag='RandomForest'):

    clf = RandomForestClassifier(n_estimators=10)
    filePath = '/Users/ming.zhou/NLP/DiscourseStructures/result/' + flag + 'Result20171115.text'

    if flag == 'ExtraTrees':
        clf = ExtraTreesClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    clf.fit(train_X, train_Y)
    Y_pred = clf.predict(test_X)

    result = open(filePath, 'a')
    resultContent = []
    resultContent.append(str(clf.score(test_X, test_Y)) + '\n')
    resultContent.append('test_Y:' + '\n')
    resultContent.append(str(test_Y) + '\n')
    resultContent.append('Y_pred:' + '\n')
    resultContent.append(str(Y_pred) + '\n')
    resultContent.append('test_Y len is ' + str(len(test_Y)) + 'Y_pred len is ' + str(len(Y_pred)) + '\n')

    print(clf.score(test_X, test_Y))
    print(accuracy_score(test_Y, Y_pred))
    print(test_Y)
    print(Y_pred)
    print('test_Y len is ', len(test_Y),'Y_pred len is ', len(Y_pred))

    compare = {}
    j = 0
    for i in range(len(test_Y)):
        YAndYpred = str(test_Y[i]) + '-' + str(Y_pred[i])
        # print(YAndYpred)
        if(test_Y[i] != Y_pred[i]):
            if YAndYpred not in compare:
                compare[YAndYpred] = 1
            else:
                compare[YAndYpred] = compare[YAndYpred] + 1

    sortedResult = sorted(compare.items(), key=lambda d: -d[1])

    resultContent.append(str(sortedResult) + '\n')
    print(sortedResult)
    result.writelines(resultContent)
    result.close()

def doTrainByCrossValidation(X, y, timesNum, flag='ExtraTrees'):
    clf = ExtraTreesClassifier(n_estimators=timesNum)
    filePath = '/Users/ming.zhou/NLP/DiscourseStructures/result/' + flag + 'CrossValidationResult20171117.text'

    result = open(filePath, 'a')
    resultContent = []
    resultContent.append('**********************step='+str(timesNum)+'**********************' + '\n')
    if flag == 'RandomForest':
        clf = RandomForestClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    # 随机划分训练集与测试集，是交叉验证中常用的函数
    sumNum = 0
    for i in range(10):
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.1, random_state=i)
        clf.fit(X_train, y_train)
        # Y_pred = clf.predict(X_test)
        score = clf.score(X_test, y_test)
        resultContent.append('index[' + str(i) + ']' + str(score) + '\n')
        print(score)

        sumNum += score
    resultContent.append('平均准确度：' + str(sumNum/10) + '\n')
    result.writelines(resultContent)
    result.close()
    print('平均准确度：', str(sumNum/10))
    return sumNum/10

# 使用交叉验证与规范化做训练
def doTrainByCVAndNorm(X, y, timesNum, flag='ExtraTrees'):
    clf = ExtraTreesClassifier(n_estimators=10)
    filePath = '/Users/ming.zhou/NLP/DiscourseStructures/result/' + flag + 'CrossValidationResult20171117.text'

    result = open(filePath, 'a')
    resultContent = []
    resultContent.append('**********************step='+str(timesNum)+'**********************' + '\n')
    if flag == 'RandomForest':
        clf = RandomForestClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    # 随机划分训练集与测试集，是交叉验证中常用的函数
    sumNum = 0
    for i in range(10):
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.1, random_state=i)
        X_train_norm = preprocessing.normalize(X_train, norm='l2')
        X_test_norm = preprocessing.normalize(X_test, norm='l2')
        clf.fit(X_train_norm, y_train)
        # Y_pred = clf.predict(X_test)
        score = clf.score(X_test_norm, y_test)
        resultContent.append('index[' + str(i) + ']' + str(score) + '\n')
        print(score)

        sumNum += score
    resultContent.append('平均准确度：' + str(sumNum/10) + '\n')
    result.writelines(resultContent)
    result.close()
    print('平均准确度：', str(sumNum/10))

def doPredict(X_train, Y_train, X_test, Y_test, timesNum, flag='ExtraTrees'):
    clf = ExtraTreesClassifier(n_estimators=timesNum)
    filePath = '/Users/ming.zhou/NLP/DiscourseStructures/result/' + flag + 'PredictResult20171117.text'

    if flag == 'RandomForest':
        clf = RandomForestClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    clf.fit(X_train, Y_train)
    Y_pred = clf.predict(X_test)
    score = clf.score(X_test, Y_test)

    # print(score)
    # print(Y_test)
    # print(Y_pred)
    # print('test_Y len is ', len(Y_test), 'Y_pred len is ', len(Y_pred))

    result = open(filePath, 'a')
    resultContent = []
    resultContent.append('**********************step=' + str(timesNum) + '**********************' + '\n')
    resultContent.append(str(score) + '\n')
    resultContent.append('test_Y:' + '\n')
    resultContent.append(str(Y_test) + '\n')
    resultContent.append('Y_pred:' + '\n')
    resultContent.append(str(Y_pred) + '\n')
    resultContent.append('test_Y len is ' + str(len(Y_test)) + '，and Y_pred len is ' + str(len(Y_pred)) + '\n')

    for i in range(len(Y_test)):
        YAndYpred = str(Y_test[i]) + '-' + str(Y_pred[i])
        # print(YAndYpred)
        if (Y_test[i] != Y_pred[i]):
            if YAndYpred not in compare:
                compare[YAndYpred] = 1
            else:
                compare[YAndYpred] = compare[YAndYpred] + 1

    # sortedResult = sorted(compare.items(), key=lambda d: -d[1])
    # print(sortedResult)
    # resultContent.append(str(sortedResult) + '\n')
    result.writelines(resultContent)
    result.close()
    return score

def doPredictByCV(X_test, Y_test, timesNum, flag='ExtraTrees'):
    clf = ExtraTreesClassifier(n_estimators=10)
    filePath = '/Users/ming.zhou/NLP/DiscourseStructures/result/' + flag + 'PredictResult20171116.text'

    if flag == 'RandomForest':
        clf = RandomForestClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    Y_pred = cross_validation.cross_val_predict(clf, X_test, Y_test, cv=20)

    # clf.fit(X_train, Y_train)
    # Y_pred = clf.predict(X_test)
    # score = clf.score(X_test, Y_test)
    score = metrics.accuracy_score(Y_test, Y_pred)

    print(score)
    print(Y_test)
    print(Y_pred)
    print('test_Y len is ', len(Y_test), 'Y_pred len is ', len(Y_pred))

    result = open(filePath, 'a')
    resultContent = []
    resultContent.append('**********************step=' + str(timesNum) + '**********************' + '\n')
    resultContent.append(str(score) + '\n')
    resultContent.append('test_Y:' + '\n')
    resultContent.append(str(Y_test) + '\n')
    resultContent.append('Y_pred:' + '\n')
    resultContent.append(str(Y_pred) + '\n')
    resultContent.append('test_Y len is ' + str(len(Y_test)) + '，and Y_pred len is ' + str(len(Y_pred)) + '\n')

    for i in range(len(Y_test)):
        YAndYpred = str(Y_test[i]) + '-' + str(Y_pred[i])
        # print(YAndYpred)
        if (Y_test[i] != Y_pred[i]):
            if YAndYpred not in compare:
                compare[YAndYpred] = 1
            else:
                compare[YAndYpred] = compare[YAndYpred] + 1

    # sortedResult = sorted(compare.items(), key=lambda d: -d[1])
    # print(sortedResult)
    # resultContent.append(str(sortedResult) + '\n')
    result.writelines(resultContent)
    result.close()
    return score

def doPredictByCVandNorm(X_train, Y_train, X_test, Y_test, timesNum, flag='ExtraTrees'):
    clf = ExtraTreesClassifier(n_estimators=10)
    filePath = '/Users/ming.zhou/NLP/DiscourseStructures/result/' + flag + 'PredictResult20171117.text'

    if flag == 'RandomForest':
        clf = RandomForestClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    normalizer = preprocessing.Normalizer().fit(X_train)
    X_train_norm = normalizer.transform(X_train)
    X_test_norm = normalizer.transform(X_test)
    clf.fit(X_train_norm, Y_train)
    Y_pred = clf.predict(X_test_norm)
    score = clf.score(X_test_norm, Y_test)

    print(score)
    print(Y_test)
    print(Y_pred)
    print('test_Y len is ', len(Y_test), 'Y_pred len is ', len(Y_pred))

    result = open(filePath, 'a')
    resultContent = []
    resultContent.append('**********************step=' + str(timesNum) + '**********************' + '\n')
    resultContent.append(str(score) + '\n')
    resultContent.append('test_Y:' + '\n')
    resultContent.append(str(Y_test) + '\n')
    resultContent.append('Y_pred:' + '\n')
    resultContent.append(str(Y_pred) + '\n')
    resultContent.append('test_Y len is ' + str(len(Y_test)) + '，and Y_pred len is ' + str(len(Y_pred)) + '\n')

    for i in range(len(Y_test)):
        YAndYpred = str(Y_test[i]) + '-' + str(Y_pred[i])
        # print(YAndYpred)
        if (Y_test[i] != Y_pred[i]):
            if YAndYpred not in compare:
                compare[YAndYpred] = 1
            else:
                compare[YAndYpred] = compare[YAndYpred] + 1

    # sortedResult = sorted(compare.items(), key=lambda d: -d[1])
    # print(sortedResult)
    # resultContent.append(str(sortedResult) + '\n')
    result.writelines(resultContent)
    result.close()
    return score

params = getFeatureVectorAndTag()
# testParams = getTestFeatureVectorAndTag()
# doTrain(params[0], params[1], 'ExtraTrees')
# doTrainByTestSet(params[0], params[1], testParams[0], testParams[1], 'ExtraTrees')
# doTrainByCrossValidation(params[0], params[1], 10, 'ExtraTrees')
# adjustParameter(params[0], params[1])
# doTrainByCVAndNorm(params[0], params[1], 2, 'ExtraTrees')
# print('*********************begin test*********************')
compare = {}

x = []
y = []
for i in range(10, 101, 5):
    x.append(i)
    scores = 0
    # for j in range(3):
    #     score = doPredict(params[0], params[1], testParams[0], testParams[1], i, 'ExtraTrees')
    #     scores += score
    score = doTrainByCrossValidation(params[0], params[1], i, 'ExtraTrees')
    y.append(score)
    print('index[' + str(i) + ']:' + str(score))

# sortedResult = sorted(compare.items(), key=lambda d: -d[1])
# print(sortedResult)
# print(scores / 30)

import pylab as pl
pl.plot(x, y)
pl.title('ExtraTrees adjust n_estimators')# give plot a title
pl.xlabel('n_estimators')# make axis labels
pl.ylabel('score')
pl.show()

# 计算程序运行总时间(秒)
elapsed = (datetime.now() - start).seconds
print('Time used : ', elapsed)